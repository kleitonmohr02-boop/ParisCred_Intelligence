"""
Módulo de Importação de Beneficiários via Excel
ParisCred Intelligence v2.0
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)

# Tentar importar openpyxl para Excel
try:
    import openpyxl
    EXCEL_ENABLED = True
except ImportError:
    EXCEL_ENABLED = False
    logger.warning("openpyxl não instalado. Para importar Excel, instale: pip install openpyxl")


class ImportadorExcel:
    """Importa dados de beneficiários de arquivos Excel"""
    
    # Mapeamento de colunas possíveis para cada campo
    CAMPOS_MAP = {
        'nome': ['nome', 'nome_completo', 'nome_cliente', 'beneficiario', 'cliente', 'nome_do_cliente', 'nome_do_beneficiario'],
        'telefone': ['telefone', 'numero', 'celular', 'phone', 'whatsapp', 'numero_telefone', 'numero_whatsapp', 'fone'],
        'cpf': ['cpf', 'documento', 'cpf_cnpj', 'cpf_cliente'],
        'email': ['email', 'e-mail', 'mail', 'email_cliente'],
        'valor': ['valor', 'valor_divida', 'valor_parcela', 'valor_emprestimo', 'limite'],
        'parcela': ['parcela', 'valor_parcela', 'parcela_atual'],
        'banco': ['banco', 'instituicao', 'orgao', 'tipo_banco'],
        'status': ['status', 'situacao', 'status_atual', 'condicao']
    }
    
    def __init__(self, arquivo_path: str):
        self.arquivo_path = arquivo_path
        self.erros = []
        self.sucessos = 0
    
    def _normalizar_telefone(self, telefone: str) -> str:
        """Normaliza número de telefone para formato padrão"""
        if not telefone:
            return ''
        
        # Remove tudo que não é número
        numeros = ''.join(filter(str.isdigit, str(telefone)))
        
        # Se começar com 0, remove
        if numeros.startswith('0'):
            numeros = numeros[1:]
        
        # Se não tiver código do país, adiciona 55 (Brasil)
        if len(numeros) == 10 or len(numeros) == 11:
            if len(numeros) == 10:
                # Telefone fixo
                numeros = '55' + numeros
            elif len(numeros) == 11:
                # Celular
                numeros = '55' + numeros
        
        return numeros
    
    def _normalizar_cpf(self, cpf: str) -> str:
        """Normaliza CPF"""
        if not cpf:
            return ''
        return ''.join(filter(str.isdigit, str(cpf)))
    
    def _identificar_coluna(self, row: Dict, campo: str) -> Optional[str]:
        """Identifica qual chave do dicionário corresponde ao campo"""
        # Normaliza as chaves do dicionário
        row_lower = {k.lower().strip(): v for k, v in row.items()}
        
        # Tenta encontrar correspondência
        for possivel in self.CAMPOS_MAP.get(campo, []):
            if possivel.lower() in row_lower:
                return row_lower[possivel.lower()]
        
        return None
    
    def _validar_telefone(self, telefone: str) -> bool:
        """Valida se telefone tem formato válido"""
        numeros = ''.join(filter(str.isdigit, telefone or ''))
        return len(numeros) >= 10 and len(numeros) <= 13
    
    def _validar_cpf(self, cpf: str) -> bool:
        """Valida CPF básico"""
        numeros = self._normalizar_cpf(cpf or '')
        return len(numeros) == 11 if numeros else True  # CPF é opcional
    
    def processar(self) -> Dict:
        """Processa o arquivo Excel e retorna lista de beneficiários"""
        if not EXCEL_ENABLED:
            return {
                'sucesso': False,
                'erro': 'openpyxl não está instalado',
                'beneficiarios': [],
                'total': 0
            }
        
        if not os.path.exists(self.arquivo_path):
            return {
                'sucesso': False,
                'erro': 'Arquivo não encontrado',
                'beneficiarios': [],
                'total': 0
            }
        
        try:
            # Abrir workbook
            wb = openpyxl.load_workbook(self.arquivo_path, data_only=True)
            ws = wb.active
            
            # Pegar cabeçalhos
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Se não tiver headers, usar índices
            if not headers or headers[0] is None:
                # Assume primeira linha como dados
                row_start = 1
                headers = [f'coluna_{i}' for i in range(ws.max_column)]
            else:
                row_start = 2
            
            # Processar linhas
            beneficiarios = []
            
            for row_num in range(row_start, ws.max_row + 1):
                row = {}
                for col_num, header in enumerate(headers, start=1):
                    value = ws.cell(row=row_num, column=col_num).value
                    row[header] = value
                
                # Extrair campos
                nome = self._identificar_coluna(row, 'nome')
                telefone = self._identificar_coluna(row, 'telefone')
                cpf = self._identificar_coluna(row, 'cpf')
                email = self._identificar_coluna(row, 'email')
                valor = self._identificar_coluna(row, 'valor')
                banco = self._identificar_coluna(row, 'banco')
                status = self._identificar_coluna(row, 'status')
                
                # Validar telefone
                telefone_normalizado = self._normalizar_telefone(telefone)
                
                if not nome:
                    self.erros.append(f"Linha {row_num}: Nome não encontrado")
                    continue
                
                if telefone and not self._validar_telefone(telefone):
                    self.erros.append(f"Linha {row_num}: Telefone inválido para {nome}")
                    continue
                
                beneficiario = {
                    'nome': str(nome).strip(),
                    'numero': telefone_normalizado,
                    'cpf': self._normalizar_cpf(cpf) if cpf else None,
                    'email': str(email).strip() if email else None,
                    'valor_divida': float(valor) if valor else None,
                    'banco': str(banco).strip() if banco else None,
                    'status': str(status).strip() if status else 'novo',
                    'importado_em': datetime.now().isoformat()
                }
                
                # Só adiciona se tiver telefone
                if telefone_normalizado:
                    beneficiarios.append(beneficiario)
                    self.sucessos += 1
                else:
                    self.erros.append(f"Linha {row_num}: Sem telefone válido para {nome}")
            
            logger.info(f"Importação concluída: {self.sucessos} успешно, {len(self.erros)} erros")
            
            return {
                'sucesso': True,
                'beneficiarios': beneficiarios,
                'total': len(beneficiarios),
                'erros': self.erros[:50]  # Limita a 50 erros
            }
            
        except Exception as e:
            logger.error(f"Erro ao processar Excel: {str(e)}")
            return {
                'sucesso': False,
                'erro': str(e),
                'beneficiarios': [],
                'total': 0
            }


class GerenciadorBeneficiarios:
    """Gerencia beneficiários no banco de dados"""
    
    @staticmethod
    def salvar_beneficiarios(beneficiarios: List[Dict]) -> Dict:
        """Salva lista de beneficiários no banco"""
        from database import Database
        
        db = Database()
        inseridos = 0
        atualizados = 0
        erros = []
        
        for benef in beneficiarios:
            try:
                with db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Verificar se já existe
                    cursor.execute(
                        "SELECT id FROM beneficiarios WHERE cpf = ? OR numero = ?",
                        (benef.get('cpf'), benef.get('numero'))
                    )
                    existente = cursor.fetchone()
                    
                    if existente:
                        # Atualizar
                        cursor.execute("""
                            UPDATE beneficiarios SET
                                nome = ?,
                                email = ?,
                                valor_divida = ?,
                                banco = ?,
                                status = ?,
                                atualizado_em = CURRENT_TIMESTAMP
                            WHERE id = ?
                        """, (
                            benef.get('nome'),
                            benef.get('email'),
                            benef.get('valor_divida'),
                            benef.get('banco'),
                            benef.get('status', 'importado'),
                            existente[0]
                        ))
                        atualizados += 1
                    else:
                        # Inserir
                        cursor.execute("""
                            INSERT INTO beneficiarios (
                                nome, cpf, numero, email,
                                valor_divida, banco, status
                            ) VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (
                            benef.get('nome'),
                            benef.get('cpf'),
                            benef.get('numero'),
                            benef.get('email'),
                            benef.get('valor_divida'),
                            benef.get('banco'),
                            benef.get('status', 'importado')
                        ))
                        inseridos += 1
                    
                    conn.commit()
                    
            except Exception as e:
                erros.append(f"Erro ao salvar {benef.get('nome')}: {str(e)}")
        
        return {
            'sucesso': True,
            'inseridos': inseridos,
            'atualizados': atualizados,
            'erros': erros
        }


# Rota de API para importação
def criar_rotas_importacao(app):
    """Cria rotas de API para importação de beneficiários"""
    from flask import request, jsonify
    
    @app.route('/api/importar/excel', methods=['POST'])
    @app.route('/api/beneficiarios/importar', methods=['POST'])
    def importar_excel():
        """Importa beneficiários de arquivo Excel"""
        if 'file' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
        
        arquivo = request.files['file']
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        # Salvar arquivo temporariamente
        import tempfile
        temp_path = os.path.join(tempfile.gettempdir(), arquivo.filename)
        arquivo.save(temp_path)
        
        try:
            # Processar
            importador = ImportadorExcel(temp_path)
            resultado = importador.processar()
            
            if resultado['sucesso']:
                # Salvar no banco
                salvar = GerenciadorBeneficiarios.salvar_beneficiarios(resultado['beneficiarios'])
                return jsonify({
                    'sucesso': True,
                    'mensagem': f"Importados {resultado['total']} beneficiários",
                    'total': resultado['total'],
                    'inseridos': salvar.get('inseridos', 0),
                    'atualizados': salvar.get('atualizados', 0),
                    'erros_importacao': resultado.get('erros', [])[:10]
                })
            else:
                return jsonify({'erro': resultado['erro']}), 400
                
        finally:
            # Limpar arquivo temporário
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    @app.route('/api/beneficiarios', methods=['GET', 'POST'])
    def api_beneficiarios():
        """API de beneficiários"""
        from database import Database
        db = Database()
        
        if request.method == 'GET':
            # Listar
            status = request.args.get('status')
            limite = request.args.get('limit', 100, type=int)
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                
                if status:
                    cursor.execute("""
                        SELECT * FROM beneficiarios
                        WHERE status = ?
                        ORDER BY criado_em DESC
                        LIMIT ?
                    """, (status, limite))
                else:
                    cursor.execute("""
                        SELECT * FROM beneficiarios
                        ORDER BY criado_em DESC
                        LIMIT ?
                    """, (limite,))
                
                results = [dict(row) for row in cursor.fetchall()]
                return jsonify(results)
        
        elif request.method == 'POST':
            # Criar
            data = request.json
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO beneficiarios (nome, cpf, numero, email, valor_divida, banco, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('nome'),
                    data.get('cpf'),
                    data.get('numero'),
                    data.get('email'),
                    data.get('valor_divida'),
                    data.get('banco'),
                    data.get('status', 'novo')
                ))
                conn.commit()
                
                return jsonify({'sucesso': True, 'id': cursor.lastrowid}), 201
    
    @app.route('/api/beneficiarios/<int:benef_id>', methods=['GET', 'PUT', 'DELETE'])
    def api_beneficiario_detalhes(benef_id):
        """Detalhes de um beneficiário"""
        from database import Database
        db = Database()
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            if request.method == 'GET':
                cursor.execute("SELECT * FROM beneficiarios WHERE id = ?", (benef_id,))
                row = cursor.fetchone()
                if not row:
                    return jsonify({'erro': 'Não encontrado'}), 404
                return jsonify(dict(row))
            
            elif request.method == 'PUT':
                data = request.json
                cursor.execute("""
                    UPDATE beneficiarios SET
                        nome = COALESCE(?, nome),
                        email = COALESCE(?, email),
                        status = COALESCE(?, status),
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (data.get('nome'), data.get('email'), data.get('status'), benef_id))
                conn.commit()
                return jsonify({'sucesso': True})
            
            elif request.method == 'DELETE':
                cursor.execute("DELETE FROM beneficiarios WHERE id = ?", (benef_id,))
                conn.commit()
                return jsonify({'sucesso': True})
    
    @app.route('/api/beneficiarios/estatisticas')
    def estatisticas_beneficiarios():
        """Estatísticas de beneficiários"""
        from database import Database
        db = Database()
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("SELECT COUNT(*) FROM beneficiarios")
            total = cursor.fetchone()[0]
            
            cursor.execute("SELECT status, COUNT(*) FROM beneficiarios GROUP BY status")
            por_status = {row[0]: row[1] for row in cursor.fetchall()}
            
            cursor.execute("SELECT banco, COUNT(*) FROM beneficiarios WHERE banco IS NOT NULL GROUP BY banco ORDER BY COUNT(*) DESC LIMIT 10")
            por_banco = {row[0]: row[1] for row in cursor.fetchall()}
            
            return jsonify({
                'total': total,
                'por_status': por_status,
                'por_banco': por_banco
            })
    
    return app